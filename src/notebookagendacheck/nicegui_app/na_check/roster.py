from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from notebookagendacheck.nicegui_app.na_check.models import RosterStudent

GRADE_KEYS = ("grade", "classgrade")
PERIOD_KEYS = ("period", "section", "block", "classperiod")
SUBJECT_KEYS = ("subject", "class", "course", "contentarea")
STUDENT_ID_KEYS = ("studentid", "student_id", "id", "sid")
NAME_KEYS = ("name", "studentname", "full_name", "fullname")
FIRST_NAME_KEYS = ("firstname", "first_name", "fname")
LAST_NAME_KEYS = ("lastname", "last_name", "lname")

SUBJECT_ALIASES = {
    "math": "Math",
    "mathematics": "Math",
    "maths": "Math",
    "science": "Science",
    "sci": "Science",
}
CSV_REQUIRED_KEYS = ("studentid", "studentname", "grade", "subject")
CSV_ALLOWED_GRADES = {"6", "7", "8"}


@dataclass(frozen=True)
class RosterValidationIssue:
    message: str
    row_number: int | None = None
    column: str | None = None


class RosterValidationError(ValueError):
    def __init__(self, message: str, issues: list[RosterValidationIssue]) -> None:
        super().__init__(message)
        self.issues = issues


def load_roster(path: Path) -> list[RosterStudent]:
    if not path.exists():
        raise FileNotFoundError(f"Roster file not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_roster_csv(path)
    if suffix == ".xlsx":
        return _load_roster_xlsx(path)
    raise ValueError(f"Unsupported roster file type: {path.suffix or '<none>'}")


def _load_roster_xlsx(path: Path) -> list[RosterStudent]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []

        headers = [normalize_header(value) for value in header_row]
        idx_grade = find_index(headers, GRADE_KEYS)
        idx_period = find_index(headers, PERIOD_KEYS)
        idx_subject = find_index(headers, SUBJECT_KEYS)
        idx_student_id = find_index(headers, STUDENT_ID_KEYS)
        idx_name = find_index(headers, NAME_KEYS)
        idx_first = find_index(headers, FIRST_NAME_KEYS)
        idx_last = find_index(headers, LAST_NAME_KEYS)

        students: list[RosterStudent] = []
        for row_num, row in enumerate(row_iter, start=2):
            values = list(row)
            if all(cell is None or str(cell).strip() == "" for cell in values):
                continue

            name = resolve_name(values, idx_name=idx_name, idx_first=idx_first, idx_last=idx_last)
            if not name:
                first_col = value_at(values, 0)
                name = first_col.strip() if first_col else f"Student {row_num - 1}"

            student_id = value_at(values, idx_student_id)
            if not student_id:
                student_id = f"ROW-{row_num - 1:04d}"

            grade = value_at(values, idx_grade) or "Unknown"
            period = value_at(values, idx_period) or "Unknown"
            subject = normalize_subject(value_at(values, idx_subject))
            students.append(
                RosterStudent(
                    grade=grade,
                    period=period,
                    subject=subject,
                    student_id=student_id,
                    student_name=name,
                )
            )
        return students
    finally:
        workbook.close()


def _load_roster_csv(path: Path) -> list[RosterStudent]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise RosterValidationError(
                "Roster CSV must include a header row.",
                issues=[RosterValidationIssue(message="CSV header row is missing.")],
            )

        normalized_map = _normalized_header_map(reader.fieldnames)
        missing_keys = [key for key in CSV_REQUIRED_KEYS if key not in normalized_map]
        if missing_keys:
            issues = [
                RosterValidationIssue(
                    message=f"Missing required column: {_display_field_name(key)}",
                    column=_display_field_name(key),
                )
                for key in missing_keys
            ]
            missing_text = ", ".join(_display_field_name(key) for key in missing_keys)
            raise RosterValidationError(
                f"Roster CSV missing required columns: {missing_text}",
                issues=issues,
            )

        issues: list[RosterValidationIssue] = []
        students: list[RosterStudent] = []
        seen_student_ids: set[str] = set()

        for row_number, row in enumerate(reader, start=2):
            if _is_csv_row_empty(row):
                continue

            student_id = normalize_cell(row.get(normalized_map["studentid"]))
            student_name = normalize_cell(row.get(normalized_map["studentname"]))
            grade_raw = normalize_cell(row.get(normalized_map["grade"]))
            subject_raw = normalize_cell(row.get(normalized_map["subject"]))
            period = normalize_cell(row.get(normalized_map["period"])) if "period" in normalized_map else "Unknown"

            if not student_id:
                issues.append(
                    RosterValidationIssue(
                        message="Missing required value for StudentID",
                        row_number=row_number,
                        column="StudentID",
                    )
                )
            if not student_name:
                issues.append(
                    RosterValidationIssue(
                        message="Missing required value for StudentName",
                        row_number=row_number,
                        column="StudentName",
                    )
                )

            normalized_grade = _normalize_csv_grade(grade_raw)
            if not normalized_grade:
                issues.append(
                    RosterValidationIssue(
                        message=f"Grade must be one of 6, 7, 8 (received: {grade_raw or '<blank>'})",
                        row_number=row_number,
                        column="Grade",
                    )
                )

            normalized_subject = normalize_subject(subject_raw)
            if not normalized_subject:
                issues.append(
                    RosterValidationIssue(
                        message=f"Subject must map to Math or Science (received: {subject_raw or '<blank>'})",
                        row_number=row_number,
                        column="Subject",
                    )
                )

            if student_id:
                if student_id in seen_student_ids:
                    issues.append(
                        RosterValidationIssue(
                            message=f"Duplicate StudentID value: {student_id}",
                            row_number=row_number,
                            column="StudentID",
                        )
                    )
                else:
                    seen_student_ids.add(student_id)

            if any(issue.row_number == row_number for issue in issues):
                continue

            students.append(
                RosterStudent(
                    grade=normalized_grade,
                    period=period or "Unknown",
                    subject=normalized_subject,
                    student_id=student_id,
                    student_name=student_name,
                )
            )

        if issues:
            raise RosterValidationError(
                f"Roster CSV validation failed with {len(issues)} issue(s).",
                issues=issues,
            )

        return students


def _normalized_header_map(headers: list[str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for header in headers:
        clean_header = normalize_header(header)
        if clean_header and clean_header not in normalized:
            normalized[clean_header] = str(header)
    return normalized


def _display_field_name(key: str) -> str:
    display = {
        "studentid": "StudentID",
        "studentname": "StudentName",
        "grade": "Grade",
        "subject": "Subject",
        "period": "Period",
    }
    return display.get(key, key)


def _is_csv_row_empty(row: dict[str, object]) -> bool:
    return all(normalize_cell(value) == "" for value in row.values())


def _normalize_csv_grade(value: str) -> str:
    cleaned = normalize_cell(value)
    if not cleaned:
        return ""

    try:
        numeric = float(cleaned)
        if numeric.is_integer():
            cleaned = str(int(numeric))
    except ValueError:
        pass

    return cleaned if cleaned in CSV_ALLOWED_GRADES else ""


def normalize_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    return "".join(ch for ch in raw if ch.isalnum())


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def value_at(values: list[object], index: int | None) -> str:
    if index is None or index < 0 or index >= len(values):
        return ""
    return normalize_cell(values[index])


def find_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def resolve_name(values: list[object], *, idx_name: int | None, idx_first: int | None, idx_last: int | None) -> str:
    direct = value_at(values, idx_name)
    if direct:
        return direct
    first = value_at(values, idx_first)
    last = value_at(values, idx_last)
    combined = " ".join(part for part in (first, last) if part)
    return combined.strip()


def normalize_subject(raw_subject: str) -> str:
    lowered = raw_subject.strip().lower()
    return SUBJECT_ALIASES.get(lowered, "")


