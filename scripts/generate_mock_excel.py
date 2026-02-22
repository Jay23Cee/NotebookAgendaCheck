from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.styles import Font


FIRST_NAMES = [
    "Aiden",
    "Alyssa",
    "Brianna",
    "Caleb",
    "Camila",
    "Carlos",
    "Daniel",
    "Delaney",
    "Elena",
    "Elijah",
    "Emma",
    "Ethan",
    "Gabriella",
    "Grayson",
    "Hannah",
    "Isaac",
    "Isabella",
    "Jasmine",
    "Jayden",
    "Jordan",
    "Kayla",
    "Leah",
    "Liam",
    "Lucas",
    "Madison",
    "Maya",
    "Mia",
    "Noah",
    "Olivia",
    "Parker",
    "Riley",
    "Samantha",
    "Sebastian",
    "Sofia",
    "Tristan",
    "Tyler",
    "Valeria",
    "William",
    "Xavier",
    "Zoe",
]

LAST_NAMES = [
    "Adams",
    "Alvarez",
    "Anderson",
    "Bailey",
    "Bennett",
    "Brooks",
    "Carter",
    "Chavez",
    "Clark",
    "Collins",
    "Diaz",
    "Edwards",
    "Flores",
    "Garcia",
    "Gonzalez",
    "Green",
    "Hall",
    "Harris",
    "Hernandez",
    "Jackson",
    "Johnson",
    "Jones",
    "Kim",
    "Lewis",
    "Lopez",
    "Martinez",
    "Miller",
    "Moore",
    "Nguyen",
    "Parker",
    "Ramirez",
    "Reed",
    "Rivera",
    "Robinson",
    "Rodriguez",
    "Sanchez",
    "Smith",
    "Taylor",
    "Thomas",
    "White",
]


def build_rows(students_per_grade: int = 54) -> list[list[str | int]]:
    random.seed(42)
    rows: list[list[str | int]] = []
    for grade in (6, 7, 8):
        for seat in range(1, students_per_grade + 1):
            first_name = random.choice(FIRST_NAMES)
            last_name = random.choice(LAST_NAMES)
            student_id = f"G{grade}-{seat:03d}"
            rows.append([student_id, first_name, last_name, grade])
    return rows


def main() -> None:
    output_file = Path("data/mock_students.xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    headers = ["StudentID", "FirstName", "LastName", "Grade"]
    sheet.append(headers)
    for row in build_rows():
        sheet.append(row)

    for idx in range(1, len(headers) + 1):
        sheet.cell(row=1, column=idx).font = Font(bold=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 8
    workbook.save(output_file)
    print(f"Created mock roster: {output_file} ({sheet.max_row - 1} students)")


if __name__ == "__main__":
    main()
