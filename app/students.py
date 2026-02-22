from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Student:
    student_id: str
    first_name: str
    last_name: str
    grade: int

    @property
    def full_name(self) -> str:
        return f"{self.last_name}, {self.first_name}"


def load_students(students_file: Path) -> list[Student]:
    if not students_file.exists():
        raise FileNotFoundError(f"Students file not found: {students_file}")

    workbook = load_workbook(students_file, data_only=True, read_only=True)
    try:
        sheet = workbook["Students"] if "Students" in workbook.sheetnames else workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        required = {"StudentID", "FirstName", "LastName", "Grade"}
        missing = required - set(header_index)
        if missing:
            raise ValueError(f"Missing required columns in sheet: {', '.join(sorted(missing))}")

        students: list[Student] = []
        for row in rows[1:]:
            if row is None:
                continue
            student_id = row[header_index["StudentID"]]
            first_name = row[header_index["FirstName"]]
            last_name = row[header_index["LastName"]]
            grade = row[header_index["Grade"]]

            if not student_id or not first_name or not last_name:
                continue

            students.append(
                Student(
                    student_id=str(student_id),
                    first_name=str(first_name),
                    last_name=str(last_name),
                    grade=int(grade),
                )
            )
        return students
    finally:
        workbook.close()


def filter_students(students: list[Student], grade: int) -> list[Student]:
    return sorted(
        [s for s in students if s.grade == grade],
        key=lambda s: (s.last_name.lower(), s.first_name.lower()),
    )
