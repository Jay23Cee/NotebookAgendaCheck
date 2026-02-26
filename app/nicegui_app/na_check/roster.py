from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.nicegui_app.na_check.models import RosterStudent

GRADE_KEYS = ("grade", "classgrade")
PERIOD_KEYS = ("period", "section", "block", "classperiod")
STUDENT_ID_KEYS = ("studentid", "student_id", "id", "sid")
NAME_KEYS = ("name", "studentname", "full_name", "fullname")
FIRST_NAME_KEYS = ("firstname", "first_name", "fname")
LAST_NAME_KEYS = ("lastname", "last_name", "lname")


def load_roster(path: Path) -> list[RosterStudent]:
    if not path.exists():
        raise FileNotFoundError(f"Roster file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            return []

        headers = [normalize_header(value) for value in header_row]
        idx_grade = find_index(headers, GRADE_KEYS)
        idx_period = find_index(headers, PERIOD_KEYS)
        idx_student_id = find_index(headers, STUDENT_ID_KEYS)
        idx_name = find_index(headers, NAME_KEYS)
        idx_first = find_index(headers, FIRST_NAME_KEYS)
        idx_last = find_index(headers, LAST_NAME_KEYS)

        students: list[RosterStudent] = []
        for row_num, row in enumerate(row_iter, start=2):
            values = list(row)
            if all(cell is None or str(cell).strip() == "" for cell in values):
                continue

            name = resolve_name(values, idx_name=idx_name, idx_first=idx_first, idx_last=idx_last)
            if not name:
                first_col = value_at(values, 0)
                name = first_col.strip() if first_col else f"Student {row_num - 1}"

            student_id = value_at(values, idx_student_id)
            if not student_id:
                student_id = f"ROW-{row_num - 1:04d}"

            grade = value_at(values, idx_grade) or "Unknown"
            period = value_at(values, idx_period) or "Unknown"
            students.append(
                RosterStudent(
                    grade=grade,
                    period=period,
                    student_id=student_id,
                    student_name=name,
                )
            )
        return students
    finally:
        workbook.close()


def normalize_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    return "".join(ch for ch in raw if ch.isalnum())


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def value_at(values: list[object], index: int | None) -> str:
    if index is None or index < 0 or index >= len(values):
        return ""
    return normalize_cell(values[index])


def find_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def resolve_name(values: list[object], *, idx_name: int | None, idx_first: int | None, idx_last: int | None) -> str:
    direct = value_at(values, idx_name)
    if direct:
        return direct
    first = value_at(values, idx_first)
    last = value_at(values, idx_last)
    combined = " ".join(part for part in (first, last) if part)
    return combined.strip()

